""" class animal:
    def __init__(self,name):
        print("constuctor1 is called")
        self.name=name
    def eat(self):
        print("eat food")
    def __del__(self):
        print("Destructor1 is called")

class cat(animal):
    def __init__(self):
        print("constuctor2 is called")
    def sound(self):
        print("Meaouuu")
    def __del__(self):
        print("Destructor2 is called")
print("===============================================")
my_cat=cat()
my_cat.eat()
my_cat.sound()
print("===============================================")
 """
""" class animal:
    def __init__(self,name):
        print("constuctor1 is called")
        self.name=name
    def eat(self):
        print("eat food")
    def __del__(self):
        print("Destructor1 is called")

class cat(animal):
    def __init__(self,name):
        print("constuctor2 is called")
        super().__init__("name")
    def sound(self):
        print("Meaouuu")
    def __del__(self):
        print("Destructor2 is called")
        super().__del__()
print("===============================================")
my_cat=cat("hero")
my_cat.eat()
my_cat.sound()
print("===============================================")

 """

""" class data:
    def __init__(self):
        print(__class__)
class point(data): 
    def __init__(self,x,y,z):
        print(__class__)
        super().__init__()

        self.x=x  
        self.y=y  
        self.z=z
p2=point(2,3,4)
print(p2.x,p2.y,p2.z)
 """
""" class point:
    def __init__(self,xcoord0=0,ycoord0=0): 
        self.xcoord0=xcoord0
        self.ycoord0=ycoord0
    def __add__(self,point_ov):
        return point(self.xcoord0+point_ov.xcoord0,self.ycoord0+point_ov.ycoord0)


point1=point(2,3)
point2=point(5,6)
point3=point1+point2

print(point3.xcoord0)
print(point3.ycoord0) """

""" f=open("test4.txt","r")
print("====================================")
#print(f.read())
print("====================================")
#print(f.read(5))
print("====================================")
#print(f.readline())
# print(f.readline())
print("====================================")
#print(f.readlines())
print("====================================")
#print(f.readlines()[0])
print("====================================")
# for line in f:
#         print(line)
print("====================================")
f.close() """
""" print("====================================")
f=open("test4_1.txt","w")
print("====================================")
f.write("embedded")
f.close()
f=open("test4_1.txt","r")
print(f.read())
print("====================================") 
 """

""" f=open("test4_1.txt","r")
#print(f.readlines())
for line in f:
    print(line)
    word=line.split()
    print(word)
    print(type(word))

 """

""" with open("test4_1.txt")as f:
    print(f.read())
 """
""" count=0
with open("test4.txt")as f:
    for line in f:
        count=count+1      
print(count)   """     
""" print("====================================")   
with open("test4.txt")as f:     
    list=f.readlines()
    print(len(list))
print("====================================")         """
""" def file_lenght(frame):
    count=0
    with open(frame)as f:
        for line in f:
            count=count+1      
    return count
print("the length of line in the file ",file_lenght("test4.txt"))  """

""" with open("test4.txt")as f:     
    list=f.read()
    print(len(list.split())) """
""" color=['red','yellow','white','black']
with open('test4_2.txt',"w")as f:
    f.write(" ".join(color))

content=open('test4_2.txt')
print(content.read()) """

""" import csv
reader=csv.reader(open('data.csv','r'))
print(reader)
for line in reader:
    print(line)
 """
""" import openpyxl
workspace=openpyxl.load_workbook("data.xlsx")
print(workspace)
for row in workspace['Sheet1'].iter_rows(values_only=True):
    print(row) """

""" import openpyxl
workbook=openpyxl.Workbook()
data=[
    ['nmae','age'],
    ['ahmed',47],
    ['mohamed',98],
    ['ali',42]
]
for row in data:
    workbook.active.append(row)
workbook.save("test4_3.xlsx")
workbook.close() """

""" import openpyxl
import openpyxl.chartsheet

workbook=openpyxl.Workbook()

new_sheet=workbook.create_sheet(title="NewSheet")

new_sheet["A1"]="M"
new_sheet["A2"]="M"
new_sheet["A3"]="M"
new_sheet["A4"]="M"
new_sheet["B4"]="M"

workbook.save("test4_4.xlsx")
workbook.close() """

""" import pandas as pd
def read_csv(file_path):
    return pd.read_csv(file_path)
def read_excel(file_path):
    return pd.read_excel(file_path)
print(read_csv("data.csv"))
print("="*50)
print(read_excel("test4_4.xlsx"))
print("="*50)
 """

""" import pandas as pd
data={
    'name':['ali','mohamed','ahmed'],
    'age':[35,65,48],
    'city':['cairo','Giza','asswan']
}

df=pd.DataFrame(data)
df.to_csv('output.csv',index=False)
df.to_excel('output.xlsx',index=False)
 """

